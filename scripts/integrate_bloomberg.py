"""
Gold trading signal — data integration (Bloomberg-primary)
==========================================================
1. Parse Bloomberg exports (workbook + COT grid + optional intermarket workbook).
2. Parse GPR (Iacoviello; not on Bloomberg).
3. Merge into unified CSVs for the signal engine.
4. Yahoo Finance runs only when explicitly enabled (--yahoo-fallback or GOLD_YF_FALLBACK=1).

Place exports under this repo (e.g. data/raw/). See references/bloomberg-data-guide.md.

**All local inputs must live inside the Gold Dashboard V2 project folder** — enforced for
``GOLD_DATA_DIR`` / ``GOLD_UPLOAD_DIR``. Yahoo Finance (optional) is the only remote source.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd

try:
    import yfinance as yf
except ImportError:
    yf = None

# Paths: project root = parent of scripts/
PROJECT_ROOT = Path(__file__).resolve().parent.parent


def _must_be_under_project(path: Path) -> Path:
    """Resolve *path* and require it to lie under PROJECT_ROOT (in-repo data only)."""
    resolved = path.resolve()
    root = PROJECT_ROOT.resolve()
    try:
        resolved.relative_to(root)
    except ValueError:
        raise ValueError(
            "GOLD_DATA_DIR and GOLD_UPLOAD_DIR must point inside the Gold Dashboard V2 project:\n"
            f"  {root}\n"
            f"Refusing: {resolved}"
        ) from None
    return resolved


def _data_path(env_key: str, default: Path) -> Path:
    """Resolve directory from env (absolute or relative to project) with project containment."""
    raw = os.environ.get(env_key)
    if raw:
        p = Path(raw)
        if not p.is_absolute():
            p = PROJECT_ROOT / p
    else:
        p = default
    return _must_be_under_project(p)


DATA_DIR = _data_path("GOLD_DATA_DIR", PROJECT_ROOT / "data")
UPLOAD_DIR = _data_path("GOLD_UPLOAD_DIR", DATA_DIR / "raw")

START = os.environ.get("GOLD_DATA_START", "2005-01-01")
END = os.environ.get("GOLD_DATA_END", "2026-12-31")

# Core macro / ETF / futures structure (Book1.xlsx Sheet2) — date/value column pairs (Excel 1-based).
BOOK1_COL_MAP = {
    "gld_shares": (6, 7),
    "gld_aum": (8, 9),
    "iau_shares": (10, 11),
    "iau_aum": (12, 13),
    "gvz": (20, 21),
    "cesiusd": (22, 23),
    "bcomgc": (24, 25),
    "cpi_yoy": (26, 27),
    "core_cpi_yoy": (28, 29),
    "nfp_change": (30, 31),
    "fed_funds": (32, 33),
    "real_yield_10y": (34, 35),
    "gc1_open_interest": (36, 37),
    "gc1_volume": (38, 39),
    "gc2_price": (40, 41),
    # Optional: append BDH pairs after column 41 for Bloomberg-primary price data (see docstring below).
    "gc1_px_open": (42, 43),
    "gc1_px_high": (44, 45),
    "gc1_px_low": (46, 47),
    "gc1_px_last": (48, 49),
    "si1_px_last": (50, 51),
    "gld_close": (52, 53),
    "gld_volume": (54, 55),
    "iau_close": (56, 57),
    "iau_volume": (58, 59),
}

# Optional second workbook: same BDH date/value layout on Sheet1 (adjust path/sheet if needed).
INTERMARKET_BOOK_PATH = "bbg_intermarket.xlsx"
INTERMARKET_SHEET = "Sheet1"
INTERMARKET_COL_MAP = {
    "DXY": (2, 3),
    "VIX": (4, 5),
    "OIL": (6, 7),
    "TIP": (8, 9),
    "TNX": (10, 11),
    "TWO": (12, 13),
    "SPX": (14, 15),
    "BTC": (16, 17),
    "USDJPY": (18, 19),
}

# Wide BDH workbook: 3-column blocks, row 3–4 headers, data from row 5 (see specs/bloomberg-bdh-paste-plan.md).
# Maps (Security, Field) -> internal bbg dict key, intermarket short name, or COT column.
WIDE_BBG_SERIES: dict[tuple[str, str], str] = {
    ("GC1 Comdty", "PX_LAST"): "gc1_px_last",
    ("GC1 Comdty", "PX_OPEN"): "gc1_px_open",
    ("GC1 Comdty", "PX_HIGH"): "gc1_px_high",
    ("GC1 Comdty", "PX_LOW"): "gc1_px_low",
    ("GC1 Comdty", "VOLUME"): "gc1_volume",
    ("GC1 Comdty", "OPEN_INT"): "gc1_open_interest",
    ("GC2 Comdty", "PX_LAST"): "gc2_price",
    ("XAUUSD Curncy", "PX_LAST"): "xau_usd",
    ("USGGT10Y Index", "PX_LAST"): "real_yield_10y",
    ("GTII10 Govt", "PX_LAST"): "tips_real_10y",
    ("FDTR Index", "PX_LAST"): "fed_funds",
    ("CPI YOY Index", "PX_LAST"): "cpi_yoy",
    ("CPUPAXFE Index", "PX_LAST"): "core_cpi_yoy",
    ("NFP TCH Index", "PX_LAST"): "nfp_change",
    ("CESIUSD Index", "PX_LAST"): "cesiusd",
    ("GLD US Equity", "PX_LAST"): "gld_close",
    ("GLD US Equity", "PX_VOLUME"): "gld_volume",
    ("GLD US Equity", "EQY_SH_OUT"): "gld_shares",
    ("GLD US Equity", "FUND_TOTAL_ASSETS"): "gld_aum",
    ("IAU US Equity", "PX_LAST"): "iau_close",
    ("IAU US Equity", "PX_VOLUME"): "iau_volume",
    ("IAU US Equity", "EQY_SH_OUT"): "iau_shares",
    ("IAU US Equity", "FUND_TOTAL_ASSETS"): "iau_aum",
    ("GVZ Index", "PX_LAST"): "gvz",
    ("BCOMGC Index", "PX_LAST"): "bcomgc",
    ("SI1 Comdty", "PX_LAST"): "si1_px_last",
    # Optional physical-market / curve overlays — verify on DES; omit from BDH if unavailable.
    ("GOLDLNPM Index", "PX_LAST"): "gold_lease_rate",
    ("GOLDIPREM Index", "PX_LAST"): "gold_india_premium",
    ("GOLDCHPREM Index", "PX_LAST"): "gold_china_premium",
    # Official sector / trade flow (monthly common) — verify DES; use custom Index if your desk publishes series in BBG.
    ("GOLDCBH Index", "PX_LAST"): "gold_cb_holdings",
    ("CHGLDIMP Index", "PX_LAST"): "gold_china_import",
    ("INGLDIMP Index", "PX_LAST"): "gold_india_import",
}

WIDE_INTERMARKET_SERIES: dict[tuple[str, str], str] = {
    ("DXY Curncy", "PX_LAST"): "DXY",
    ("USDJPY Curncy", "PX_LAST"): "USDJPY",
    ("VIX Index", "PX_LAST"): "VIX",
    ("SPX Index", "PX_LAST"): "SPX",
    ("CL1 Comdty", "PX_LAST"): "OIL",
    ("XBTUSD Curncy", "PX_LAST"): "BTC",
    ("TIP US Equity", "PX_LAST"): "TIP",
    ("USGG10YR Index", "PX_LAST"): "TNX",
    ("USGG2YR Index", "PX_LAST"): "USGG2YR",
    ("USGGBE10 Index", "PX_LAST"): "USGGBE10",
    ("USGG3M Index", "PX_LAST"): "TWO",
    ("H0A0 Index", "PX_LAST"): "HY_OAS",
}

WIDE_COT_SERIES: dict[tuple[str, str], str] = {
    ("CFFDUMML Index", "PX_LAST"): "managed_money_long",
    ("CFFDUMMS Index", "PX_LAST"): "managed_money_short",
    ("CFFDUMMN Index", "PX_LAST"): "managed_money_net",
    ("CFFDUPML Index", "PX_LAST"): "producer_long",
    ("CFFDUPMS Index", "PX_LAST"): "producer_short",
    ("CFFDUPMN Index", "PX_LAST"): "producer_net",
    ("CFFDUSWN Index", "PX_LAST"): "swap_dealers_net",
    # Extended COMEX gold positioning (weekly BDH "Per","W") — confirm on DES / COT <GO>.
    ("CFFDUORN Index", "PX_LAST"): "other_reportables_net",
    ("GCNCN Index", "PX_LAST"): "legacy_noncomm_net",
}


def _parse_args():
    p = argparse.ArgumentParser(description="Merge Bloomberg exports (+ optional Yahoo fallback) into data/*.csv")
    p.add_argument(
        "--yahoo-fallback",
        action="store_true",
        help="Download missing series from Yahoo Finance after Bloomberg parse",
    )
    return p.parse_args()


def yahoo_fallback_enabled(cli_flag: bool) -> bool:
    if cli_flag:
        return True
    v = os.environ.get("GOLD_YF_FALLBACK", "").strip().lower()
    return v in ("1", "true", "yes", "on")


def _coerce_excel_date(val: object) -> pd.Timestamp | None:
    """Normalize openpyxl cell values (datetime or Excel serial) to pandas Timestamp."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    if isinstance(val, date) and not isinstance(val, datetime):
        return pd.Timestamp(datetime(val.year, val.month, val.day))
    if isinstance(val, (int, float)):
        ts = pd.to_datetime(float(val), unit="D", origin="1899-12-30", errors="coerce")
        if pd.isna(ts):
            return None
        return pd.Timestamp(ts)
    try:
        ts = pd.to_datetime(val, errors="coerce")
    except (TypeError, ValueError):
        return None
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts)


def _read_wide_block_series(ws, date_col: int, val_col: int, *, data_start_row: int = 5) -> pd.Series | None:
    """Read one Date | Value block from a wide BDH sheet (1-based Excel columns)."""
    dates: list[pd.Timestamp] = []
    values: list[float] = []
    for row in range(data_start_row, ws.max_row + 1):
        raw_d = ws.cell(row=row, column=date_col).value
        raw_v = ws.cell(row=row, column=val_col).value
        d = _coerce_excel_date(raw_d)
        if d is None or raw_v is None:
            continue
        if isinstance(raw_v, str) and raw_v.startswith("#"):
            continue
        try:
            values.append(float(raw_v))
        except (TypeError, ValueError):
            continue
        dates.append(d)
    if not dates:
        return None
    s = pd.Series(values, index=pd.DatetimeIndex(dates))
    s = s[~s.index.duplicated(keep="first")].sort_index()
    return s


def _resolve_bdh_export_path() -> Path | None:
    env = os.environ.get("GOLD_BDH_EXPORT", "").strip()
    if env:
        p = Path(env)
        if not p.is_absolute():
            p = PROJECT_ROOT / p
        if not p.is_file():
            print(f"  WARNING: GOLD_BDH_EXPORT not found: {p}")
            return None
        return _must_be_under_project(p)
    bloom = UPLOAD_DIR / "Bloomberg"
    if not bloom.is_dir():
        return None
    cands = [p for p in bloom.glob("bbg_bdh_export*.xlsx") if not p.name.startswith("~$")]
    if not cands:
        return None
    cands.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return cands[0]


def parse_wide_bdh_export() -> tuple[dict[str, pd.Series], dict[str, pd.Series], pd.DataFrame]:
    """
    Parse `data/raw/Bloomberg/bbg_bdh_export*.xlsx` (3-col BDH blocks, data from row 5).
    Returns (bbg_series, intermarket_series, cot_frame).
    """
    path = _resolve_bdh_export_path()
    if path is None or not path.is_file():
        print("\n[BLOOMBERG] Wide BDH export: not found (optional). Set GOLD_BDH_EXPORT or add data/raw/Bloomberg/bbg_bdh_export*.xlsx")
        return {}, {}, pd.DataFrame()

    print(f"\n[BLOOMBERG] Parsing wide BDH workbook: {path.name}")
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "BBG_BDH_Excel_Paste_Table" in wb.sheetnames:
        ws = wb["BBG_BDH_Excel_Paste_Table"]
    elif wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
        print(f"  Using sheet: {ws.title!r}")
    else:
        print("  SKIP: workbook has no sheets")
        wb.close()
        return {}, {}, pd.DataFrame()

    bbg: dict[str, pd.Series] = {}
    im: dict[str, pd.Series] = {}
    cot_cols: dict[str, pd.Series] = {}

    def _ingest_block(key: tuple[str, str], date_col: int, val_col: int, data_start_row: int) -> None:
        s = _read_wide_block_series(ws, date_col, val_col, data_start_row=data_start_row)
        if s is None or s.empty:
            return
        if key in WIDE_BBG_SERIES:
            name = WIDE_BBG_SERIES[key]
            s.name = name
            bbg[name] = s
            print(f"  {name}: {len(s)} rows, {s.index.min().date()} to {s.index.max().date()}")
        elif key in WIDE_INTERMARKET_SERIES:
            name = WIDE_INTERMARKET_SERIES[key]
            s.name = name
            im[name] = s
            print(f"  [IM] {name}: {len(s)} rows")
        elif key in WIDE_COT_SERIES:
            name = WIDE_COT_SERIES[key]
            s.name = name
            cot_cols[name] = s
            print(f"  [COT] {name}: {len(s)} rows")

    # Preferred: 3-column blocks — Security / Field on row 4; BDH spill from row 5 (see specs/bloomberg-bdh-paste-plan.md).
    for c in range(1, ws.max_column + 1, 3):
        sec = ws.cell(row=4, column=c).value
        fld = ws.cell(row=4, column=c + 1).value
        if sec is None or fld is None:
            continue
        if not isinstance(sec, str) or not isinstance(fld, str):
            continue
        key = (sec.strip(), fld.strip())
        _ingest_block(key, c, c + 1, data_start_row=5)

    # Compact export: Security row 1, Field row 2, date|value from row 4 (no spacer column).
    for c in range(4, ws.max_column + 1, 2):
        sec = ws.cell(row=1, column=c).value
        fld = ws.cell(row=2, column=c).value
        if not isinstance(sec, str) or not isinstance(fld, str):
            continue
        if not sec.strip():
            continue
        key = (sec.strip(), fld.strip())
        if key not in WIDE_BBG_SERIES and key not in WIDE_INTERMARKET_SERIES and key not in WIDE_COT_SERIES:
            continue
        _ingest_block(key, c, c + 1, data_start_row=4)

    cot_df = pd.DataFrame(cot_cols) if cot_cols else pd.DataFrame()
    wb.close()
    return bbg, im, cot_df


def parse_sheet_date_value_pairs(ws, col_map: dict[str, tuple[int, int]]) -> dict[str, pd.Series]:
    """Read BDH-style blocks: for each logical name, (date_col, value_col) are 1-based Excel columns."""
    out: dict[str, pd.Series] = {}
    for name, (date_col, val_col) in col_map.items():
        dates: list = []
        values: list = []
        for row in range(4, ws.max_row + 1):
            d = ws.cell(row=row, column=date_col).value
            v = ws.cell(row=row, column=val_col).value
            if d is None or v is None:
                continue
            if isinstance(v, str) and v.startswith("#"):
                continue
            try:
                dates.append(pd.to_datetime(d))
                values.append(float(v))
            except (ValueError, TypeError):
                continue
        if dates:
            s = pd.Series(values, index=pd.DatetimeIndex(dates), name=name)
            s = s[~s.index.duplicated(keep="first")].sort_index()
            out[name] = s
            print(f"  {name}: {len(s)} rows, {s.index.min().date()} to {s.index.max().date()}")
    return out


def parse_bloomberg_book1() -> dict[str, pd.Series]:
    """Parse Bloomberg data from Book1.xlsx Sheet2."""
    path = UPLOAD_DIR / "Book1.xlsx"
    print("\n[BLOOMBERG] Parsing Book1.xlsx Sheet2...")
    if not path.is_file():
        print(f"  SKIP: not found: {path}")
        return {}
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Sheet2" not in wb.sheetnames:
        print("  SKIP: Sheet2 missing")
        return {}
    ws = wb["Sheet2"]
    series_data = parse_sheet_date_value_pairs(ws, BOOK1_COL_MAP)
    missing_core = [k for k in ("gld_shares", "real_yield_10y", "fed_funds", "cesiusd") if k not in series_data]
    if missing_core:
        print(f"  Note: missing some core Book1 series: {missing_core}")
    return series_data


def parse_bloomberg_intermarket_workbook() -> dict[str, pd.Series]:
    """Optional intermarket BDH export (DXY, VIX, crude, etc.)."""
    path = UPLOAD_DIR / INTERMARKET_BOOK_PATH
    print(f"\n[BLOOMBERG] Parsing {INTERMARKET_BOOK_PATH} {INTERMARKET_SHEET}...")
    if not path.is_file():
        print(f"  SKIP: not found: {path}")
        return {}
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if INTERMARKET_SHEET not in wb.sheetnames:
        print(f"  SKIP: sheet {INTERMARKET_SHEET!r} missing")
        return {}
    ws = wb[INTERMARKET_SHEET]
    return parse_sheet_date_value_pairs(ws, INTERMARKET_COL_MAP)


def parse_cot_data() -> pd.DataFrame:
    """Parse Bloomberg COT data from grid1.xlsx."""
    path = UPLOAD_DIR / "grid1.xlsx"
    print("\n[BLOOMBERG] Parsing COT data from grid1.xlsx...")
    if not path.is_file():
        print(f"  SKIP: not found: {path}")
        return pd.DataFrame()
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    cot_cols = {
        "managed_money_long": 13,
        "managed_money_short": 15,
        "managed_money_net": 17,
        "producer_long": 19,
        "producer_short": 21,
        "producer_net": 23,
        "swap_dealers_net": 25,
    }
    cot_data: dict[str, pd.Series] = {}
    for name, val_col in cot_cols.items():
        dates = []
        values = []
        for row in range(4, ws.max_row + 1):
            d = ws.cell(row=row, column=12).value
            v = ws.cell(row=row, column=val_col).value
            if d is None or v is None:
                continue
            try:
                dates.append(pd.to_datetime(d))
                values.append(float(v))
            except (ValueError, TypeError):
                continue
        if dates:
            s = pd.Series(values, index=pd.DatetimeIndex(dates), name=name)
            s = s[~s.index.duplicated(keep="first")].sort_index()
            cot_data[name] = s
            print(f"  {name}: {len(s)} rows, {s.index.min().date()} to {s.index.max().date()}")
    return pd.DataFrame(cot_data) if cot_data else pd.DataFrame()


def parse_gpr_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse GPR from Iacoviello files under data/raw/ (.dta preferred, then .xls)."""
    print("\n[GPR] Parsing GPR Index data...")
    gpr_daily = pd.DataFrame()
    for name in ("data_gpr_daily_recent.dta", "data_gpr_daily_recent.xls"):
        daily_path = UPLOAD_DIR / name
        if not daily_path.is_file():
            continue
        if daily_path.suffix.lower() == ".dta":
            gpr_daily = pd.read_stata(daily_path)
        else:
            gpr_daily = pd.read_excel(daily_path)
        gpr_daily["date"] = pd.to_datetime(gpr_daily["date"])
        gpr_daily = gpr_daily.set_index("date")
        gpr_daily = gpr_daily[["GPRD", "GPRD_ACT", "GPRD_THREAT"]].copy()
        gpr_daily.columns = ["gpr_index", "gpr_acts", "gpr_threats"]
        gpr_daily = gpr_daily[~gpr_daily.index.duplicated(keep="first")].sort_index()
        print(f"  GPR Daily ({name}): {len(gpr_daily)} rows, {gpr_daily.index.min().date()} to {gpr_daily.index.max().date()}")
        break
    else:
        print(f"  SKIP daily: place data_gpr_daily_recent.dta or .xls in {UPLOAD_DIR}")

    gpr_monthly = pd.DataFrame()
    for name in ("data_gpr_export.dta", "data_gpr_export.xls"):
        monthly_path = UPLOAD_DIR / name
        if not monthly_path.is_file():
            continue
        if monthly_path.suffix.lower() == ".dta":
            gpr_monthly = pd.read_stata(monthly_path)
        else:
            gpr_monthly = pd.read_excel(monthly_path)
        gpr_monthly["month"] = pd.to_datetime(gpr_monthly["month"])
        gpr_monthly = gpr_monthly.set_index("month")
        gpr_monthly = gpr_monthly[["GPR", "GPRT", "GPRA"]].copy()
        gpr_monthly.columns = ["gpr_monthly", "gpr_threats_monthly", "gpr_acts_monthly"]
        gpr_monthly = gpr_monthly[gpr_monthly.index > pd.Timestamp("1950-01-01")]
        print(
            f"  GPR Monthly ({name}): {len(gpr_monthly)} rows, "
            f"{gpr_monthly.index.min().date()} to {gpr_monthly.index.max().date()}"
        )
        break
    else:
        print(f"  SKIP monthly: place data_gpr_export.dta or .xls in {UPLOAD_DIR}")

    return gpr_daily, gpr_monthly


def download_yf(ticker: str, name: str) -> pd.DataFrame | None:
    if yf is None:
        print("  ERROR: yfinance not installed; pip install yfinance")
        return None
    print(f"  Downloading {name} ({ticker})...")
    try:
        df = yf.download(ticker, start=START, end=END, progress=False)
        if df.empty:
            print(f"  WARNING: No data for {ticker}")
            return None
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.index = pd.to_datetime(df.index)
        df.index = df.index.tz_localize(None) if df.index.tz else df.index
        print(f"  Got {len(df)} rows for {name} ({df.index.min().date()} to {df.index.max().date()})")
        return df
    except Exception as e:
        print(f"  ERROR downloading {ticker}: {e}")
        return None


def gold_ohlcv_from_bbg(bbg: dict[str, pd.Series]) -> pd.DataFrame | None:
    """Build gold_price-like DataFrame from Bloomberg GC1 fields on Book1."""
    if "gc1_px_last" not in bbg:
        return None
    close = bbg["gc1_px_last"].dropna()
    if close.empty:
        return None
    idx = close.index
    open_ = bbg.get("gc1_px_open", close).reindex(idx)
    high = bbg.get("gc1_px_high", close).reindex(idx)
    low = bbg.get("gc1_px_low", close).reindex(idx)
    vol = bbg.get("gc1_volume", pd.Series(np.nan, index=idx)).reindex(idx)
    out = pd.DataFrame({"Open": open_, "High": high, "Low": low, "Close": close, "Volume": vol}, index=idx)
    out = out.sort_index()
    print(f"\n[BBG] Built gold OHLCV from GC1 fields: {len(out)} rows")
    return out


def silver_from_bbg(bbg: dict[str, pd.Series]) -> pd.DataFrame | None:
    if "si1_px_last" not in bbg:
        return None
    s = bbg["si1_px_last"].dropna()
    if s.empty:
        return None
    df = pd.DataFrame({"Close": s})
    print(f"\n[BBG] Built silver from SI1: {len(df)} rows")
    return df


def etf_from_bbg(bbg: dict[str, pd.Series], prefix: str) -> pd.DataFrame | None:
    close_k = f"{prefix}_close"
    vol_k = f"{prefix}_volume"
    if close_k not in bbg:
        return None
    close = bbg[close_k].dropna()
    if close.empty:
        return None
    idx = close.index
    vol = bbg.get(vol_k, pd.Series(np.nan, index=idx)).reindex(idx)
    df = pd.DataFrame({"Close": close, "Volume": vol}, index=idx).sort_index()
    print(f"\n[BBG] Built {prefix.upper()} ETF OHLCV from Bloomberg: {len(df)} rows")
    return df


def apply_bbg_to_intermarket(intermarket_data: dict[str, pd.Series], bbg: dict[str, pd.Series]) -> None:
    """Bloomberg macro columns override / fill intermarket frame."""
    mapping = [
        ("real_yield_10y", "REAL_YIELD_10Y"),
        ("fed_funds", "FED_FUNDS"),
        ("cesiusd", "ECON_SURPRISE"),
        ("gvz", "GVZ"),
        ("cpi_yoy", "CPI_YOY"),
        ("core_cpi_yoy", "CORE_CPI_YOY"),
        ("nfp_change", "NFP_CHANGE"),
        ("bcomgc", "BCOMGC"),
        ("gold_lease_rate", "GOLD_LEASE"),
        ("gold_india_premium", "GOLD_INDIA_PREM"),
        ("gold_china_premium", "GOLD_CHINA_PREM"),
        ("gold_cb_holdings", "GOLD_CB_HOLDINGS"),
        ("gold_china_import", "GOLD_CHINA_IMPORT"),
        ("gold_india_import", "GOLD_INDIA_IMPORT"),
    ]
    for src, dst in mapping:
        if src in bbg:
            intermarket_data[dst] = bbg[src]


def yahoo_fill_prices_and_intermarket(intermarket_data: dict[str, pd.Series]) -> tuple[pd.DataFrame | None, pd.DataFrame | None, pd.DataFrame | None, pd.DataFrame | None]:
    """Download Yahoo series used by the legacy pipeline."""
    gold = download_yf("GC=F", "Gold Futures")
    silver = download_yf("SI=F", "Silver Futures")
    tickers = {
        "DXY": "DX-Y.NYB",
        "TIP": "TIP",
        "TNX": "^TNX",
        "TWO": "^IRX",
        "VIX": "^VIX",
        "SPX": "^GSPC",
        "OIL": "CL=F",
        "BTC": "BTC-USD",
        "USDJPY": "JPY=X",
    }
    for name, ticker in tickers.items():
        if name in intermarket_data:
            continue
        df = download_yf(ticker, name)
        if df is not None and "Close" in df.columns:
            intermarket_data[name] = df["Close"]
    gld_yf = download_yf("GLD", "SPDR Gold Shares")
    iau_yf = download_yf("IAU", "iShares Gold Trust")
    return gold, silver, gld_yf, iau_yf


def main() -> None:
    args = _parse_args()
    use_yahoo = yahoo_fallback_enabled(args.yahoo_fallback)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 70)
    print("GOLD SIGNAL — DATA INTEGRATION (Bloomberg-primary)")
    print(f"Data dir: {DATA_DIR}")
    print(f"Upload dir: {UPLOAD_DIR}")
    print(f"Yahoo fallback: {'ON' if use_yahoo else 'OFF'} (set GOLD_YF_FALLBACK=1 or --yahoo-fallback to enable)")
    print("=" * 70)

    # --- Step 1: Bloomberg ---
    print("\n" + "=" * 70)
    print("STEP 1: Parse Bloomberg exports")
    print("=" * 70)
    bbg_wide, im_wide, cot_wide = parse_wide_bdh_export()
    bbg_book = parse_bloomberg_book1()
    bbg = {**bbg_wide, **bbg_book}
    bbg_im = parse_bloomberg_intermarket_workbook()
    cot_df = parse_cot_data()
    if cot_df.empty and not cot_wide.empty:
        cot_df = cot_wide

    intermarket_data: dict[str, pd.Series] = dict(im_wide)
    for k, ser in bbg_im.items():
        intermarket_data[k] = ser
    apply_bbg_to_intermarket(intermarket_data, bbg)

    gold = gold_ohlcv_from_bbg(bbg)
    silver = silver_from_bbg(bbg)
    gld_yf = etf_from_bbg(bbg, "gld")
    iau_yf = etf_from_bbg(bbg, "iau")

    # --- Step 2: GPR ---
    print("\n" + "=" * 70)
    print("STEP 2: Parse GPR (public)")
    print("=" * 70)
    gpr_daily, gpr_monthly = parse_gpr_data()

    # --- Step 3: Optional Yahoo fallback ---
    print("\n" + "=" * 70)
    print("STEP 3: Yahoo Finance (fallback only)")
    print("=" * 70)
    if use_yahoo:
        if yf is None:
            print("  yfinance not installed — cannot run fallback. pip install yfinance")
        else:
            yg, ys, ygld, yiau = yahoo_fill_prices_and_intermarket(intermarket_data)
            if gold is None:
                gold = yg
            if silver is None:
                silver = ys
            if gld_yf is None:
                gld_yf = ygld
            if iau_yf is None:
                iau_yf = yiau
            apply_bbg_to_intermarket(intermarket_data, bbg)
    else:
        print("  Skipped (Yahoo fallback disabled).")

    # --- Step 4: Merge / validate / save ---
    print("\n" + "=" * 70)
    print("STEP 4: Merge, validate, write CSVs")
    print("=" * 70)

    if gold is None:
        print("\nERROR: No gold price series. Either:")
        print("  - Add data/raw/Bloomberg/bbg_bdh_export*.xlsx (wide BDH) with GC1 PX_LAST, or")
        print("  - Paste GC1 PX_LAST into Book1.xlsx Sheet2 (see BOOK1_COL_MAP), or")
        print("  - Run with --yahoo-fallback for Yahoo Finance (optional).")
        sys.exit(1)

    if gld_yf is None:
        print(
            "\nWARNING: No GLD ETF data (needed for sentiment / premium). "
            "Add gld_close/gld_volume columns to Book1 or use --yahoo-fallback."
        )

    gold.to_csv(DATA_DIR / "gold_price.csv")
    if "xau_usd" in bbg:
        pd.DataFrame({"Close": bbg["xau_usd"].sort_index()}).to_csv(DATA_DIR / "xauusd_spot.csv")
        print(f"  Saved xauusd_spot (XAUUSD COB / execution spine): {len(bbg['xau_usd'])} rows")
    if silver is not None:
        silver.to_csv(DATA_DIR / "silver_price.csv")
    if gld_yf is not None:
        gld_yf.to_csv(DATA_DIR / "gld_etf.csv")
    if iau_yf is not None:
        iau_yf.to_csv(DATA_DIR / "iau_etf.csv")

    if intermarket_data:
        im_df = pd.DataFrame(intermarket_data)
        if "tips_real_10y" in bbg:
            im_df["TIPS_REAL_10Y"] = bbg["tips_real_10y"].reindex(im_df.index)
        im_df.to_csv(DATA_DIR / "intermarket.csv")
        print(f"  Saved intermarket: {im_df.shape}, {im_df.index.min()} to {im_df.index.max()}")
    else:
        print("  WARNING: intermarket empty — dimension 2/3/4/7 will be degraded")

    if not cot_df.empty:
        cot_df.to_csv(DATA_DIR / "cot_data.csv")
        print(f"  Saved COT: {cot_df.shape}")
    else:
        print("  WARNING: no COT data — dimension 5 degraded")

    if not gpr_daily.empty:
        gpr_daily.to_csv(DATA_DIR / "gpr_daily.csv")
        print(f"  Saved GPR daily: {gpr_daily.shape}")
    else:
        print("  WARNING: no GPR daily — dimension 7 degraded")
    if not gpr_monthly.empty:
        gpr_monthly.to_csv(DATA_DIR / "gpr_monthly.csv")

    etf_fund = pd.DataFrame()
    for name in ["gld_shares", "gld_aum", "iau_shares", "iau_aum"]:
        if name in bbg:
            etf_fund[name] = bbg[name]
    if not etf_fund.empty:
        etf_fund.to_csv(DATA_DIR / "etf_fundamentals.csv")
        print(f"  Saved ETF fundamentals: {etf_fund.shape}")

    mkt_struct = pd.DataFrame()
    for name in ["gc1_open_interest", "gc1_volume", "gc2_price"]:
        if name in bbg:
            mkt_struct[name] = bbg[name]
    if not mkt_struct.empty:
        mkt_struct.to_csv(DATA_DIR / "market_structure_bbg.csv")
        print(f"  Saved market structure (BBG): {mkt_struct.shape}")

    gold_close = gold["Close"].dropna()
    season_df = pd.DataFrame(index=gold_close.index)
    season_df["month"] = season_df.index.month
    season_df["day_of_week"] = season_df.index.dayofweek
    season_df["is_indian_wedding_season"] = season_df["month"].isin([11, 12, 1, 2, 3]).astype(int)
    season_df["is_chinese_ny_period"] = ((season_df["month"] == 1) | (season_df["month"] == 2)).astype(int)
    season_df["is_september"] = (season_df["month"] == 9).astype(int)
    season_df["quarter"] = season_df.index.quarter
    season_df.to_csv(DATA_DIR / "seasonality.csv")
    print(f"  Saved seasonality: {season_df.shape}")

    geo_df = pd.DataFrame(index=gold_close.index)
    geo_df["gold_realized_vol_20d"] = gold_close.pct_change().rolling(20).std() * np.sqrt(252)
    geo_df["gold_realized_vol_5d"] = gold_close.pct_change().rolling(5).std() * np.sqrt(252)
    geo_df["vol_of_vol"] = geo_df["gold_realized_vol_20d"].rolling(20).std()
    geo_df.to_csv(DATA_DIR / "geopolitical_proxy.csv")
    print(f"  Saved geopolitical proxy: {geo_df.shape}")

    print("\n" + "=" * 70)
    print("DATA INTEGRATION COMPLETE — SUMMARY")
    print("=" * 70)
    files = sorted(f for f in os.listdir(DATA_DIR) if f.endswith(".csv"))
    summary = {}
    for f in files:
        df = pd.read_csv(DATA_DIR / f, index_col=0, parse_dates=True)
        info = {
            "rows": len(df),
            "columns": list(df.columns)[:12],
            "date_range": f"{df.index.min().date()} to {df.index.max().date()}" if len(df) > 0 else "empty",
        }
        summary[f] = info
        print(f"  {f}: {info['rows']} rows, {info['date_range']}")

    with open(DATA_DIR / "data_summary.json", "w") as fp:
        json.dump(summary, fp, indent=2, default=str)

    print("\nAll data ready for signal generation.")


if __name__ == "__main__":
    main()
